import numpy as np
from openpyxl import load_workbook

wb = load_workbook("./data/donnees optim sujet interco.xlsx", data_only=True)
sheetNord = wb[wb.sheetnames[1]]
sheetSud = wb[wb.sheetnames[2]]

# ICI le scénario qu'on veut tester, à modifier !
scenario = 2
# Scénario 0 : n premières heures
# Scénario 1 : mois de mai
# Scénario 2 : mois de décembre

if scenario == 0:
    # modifier le nombre d'heures à prendre si on veut tester les n premières heures
    nombreHeuresATester = len(sheetNord["B"])  # toute l'année
    nombreHeuresATester = 24 * 30  # 1 mois

    debut = 1
    fin = debut + nombreHeuresATester - 1

if scenario == 1:
    # MOIS DE MAI
    debut = 2877
    fin = 3617

if scenario == 2:
    # MOIS DE DECEMBRE
    debut = 8013
    fin = 8760


def cleanValue(value):
    if value < 0:
        return 0
    return float(value)


dates = [cell.value for cell in sheetNord["A"][debut:fin]]

consoNord = np.array([cleanValue(cell.value) for cell in sheetNord["B"][debut:fin]])
prodSolaireNord = [cleanValue(cell.value) for cell in sheetNord["C"][debut:fin]]

consoSud = np.array([cleanValue(cell.value) for cell in sheetSud["B"][debut:fin]])
prodSolaireSud = [cleanValue(cell.value) for cell in sheetSud["C"][debut:fin]]
prodEolienSud = [cleanValue(cell.value) for cell in sheetSud["D"][debut:fin]]
prodHydroSud = [cleanValue(cell.value) for cell in sheetSud["E"][debut:fin]]
prodBioenergiesSud = [cleanValue(cell.value) for cell in sheetSud["F"][debut:fin]]

nbHeures = len(consoNord)

assert (
    len(consoNord)
    == len(prodSolaireNord)
    == len(consoSud)
    == len(prodSolaireSud)
    == len(prodEolienSud)
    == len(prodHydroSud)
    == len(prodBioenergiesSud)
)

""" Quelques données en plus """
# Interconnexion
capaciteIntercoInitiale = 100  # MW, valable dans les deux sens
