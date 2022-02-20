from openpyxl import Workbook
from datetime import datetime

from readExcel import dates


def writeResultsInExcel(mesZones, nbHeures):
    wb = Workbook()

    """ On note l'interconnexion """
    ws1 = wb.active
    ws1.title = "Interconnexion"

    ws1.append(["Dates"] + [f"vers {zone.nom.name}" for zone in mesZones.values()])
    for h in range(nbHeures):
        ws1.append(
            [dates[h]] + [zone.solutionIntercoVersMoi[h] for zone in mesZones.values()]
        )

    for zone in mesZones.values():
        """On note le dispatchable"""
        ws = wb.create_sheet(title=f"{zone.nom.name}_dispatchable")
        # Noms des centrales
        ws.append(
            ["Dates"] + [prod.nomCentrale for prod in zone.producteursDispatchable]
        )
        ws.append([""] + [prod.type.name for prod in zone.producteursDispatchable])
        # Données solutions
        for h in range(nbHeures):
            ws.append(
                [dates[h]]
                + [prod.solutionProduction[h] for prod in zone.producteursDispatchable]
            )

        """On note le fatal"""
        ws = wb.create_sheet(title=f"{zone.nom.name}_fatal")
        # Noms des centrales
        ws.append(["Dates"] + [prod.nomCentrale for prod in zone.producteursFatal])
        # Données solutions
        for h in range(nbHeures):
            ws.append(
                [dates[h]] + [prod.production[h] for prod in zone.producteursFatal]
            )

    timestamp = datetime.now().strftime("%d%m%Y-%H%M%S")
    dest_filename = f"src/postTraitementFiles/results_{timestamp}.xlsx"
    wb.save(filename=dest_filename)
