import numpy as np
from openpyxl import load_workbook

wb = load_workbook("./data/donnees optim sujet interco.xlsx")
sheetNord = wb[wb.sheetnames[1]]
sheetSud = wb[wb.sheetnames[2]]


# nbHeures = len(sheetNord["B"])
nbHeures = 24 * 30  # 1 mois
# nbHeures = 24

def cleanValue(value):
    if value < 0:
        return 0
    return float(value)


consoNord = np.array([cleanValue(cell.value) for cell in sheetNord["B"][1:nbHeures]])
prodSolaireNord = [cleanValue(cell.value) for cell in sheetNord["C"][1:nbHeures]]

consoSud = np.array([cleanValue(cell.value) for cell in sheetSud["B"][1:nbHeures]])
prodSolaireSud = [cleanValue(cell.value) for cell in sheetSud["C"][1:nbHeures]]
prodEolienSud = [cleanValue(cell.value) for cell in sheetSud["D"][1:nbHeures]]
prodHydroSud = [cleanValue(cell.value) for cell in sheetSud["E"][1:nbHeures]]
prodBioenergiesSud = [cleanValue(cell.value) for cell in sheetSud["F"][1:nbHeures]]

nbHeures = len(consoNord)

assert (
    len(consoNord)
    == len(prodSolaireNord)
    == len(consoSud)
    == len(prodSolaireSud)
    == len(prodEolienSud)
    == len(prodHydroSud)
    == len(prodBioenergiesSud)
)

""" Quelques données en plus """
# Interconnexion
capaciteIntercoInitiale = 100  # MW, valable dans les deux sens
